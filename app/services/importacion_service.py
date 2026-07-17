"""
Servicio de importación de catálogo (Fase 11): parseo de Excel/XML,
emparejamiento por nombre contra el catálogo existente (sugerencia, nunca
automática), y confirmación (crear/vincular productos + ingreso opcional).

Todo lo que se crea o toca en la BD real pasa primero por una fila en
`importacion_linea` con `decision='pendiente'` — el dueño decide en la
vista previa antes de que `confirmar_lote` haga cualquier escritura.
"""
import difflib
import re
import unicodedata
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
from io import BytesIO
from typing import Optional

import openpyxl
from fastapi import HTTPException
from sqlalchemy import select

from app.core.constants import (
    IMPORT_TIPO_EXCEL,
    IMPORT_TIPO_XML,
    IMPORT_TIPO_PDF,
    IMPORT_ESTADO_REVISION,
    IMPORT_ESTADO_CONFIRMADO,
    IMPORT_ESTADO_CANCELADO,
    IMPORT_DECISION_CREAR,
    IMPORT_DECISION_VINCULAR,
    IMPORT_DECISION_IGNORAR,
    IMPORT_DECISIONES,
    MOV_COMPRA,
)
from app.core.database import database
from app.models import (
    marca,
    categoria,
    producto,
    inventario,
    ingreso_inventario,
    importacion_lote,
    importacion_linea,
)
from app.services.inventario_service import registrar_movimiento
from app.services.historial_precio_service import registrar_cambio_precio

CFDI_NS = {"cfdi": "http://www.sat.gob.mx/cfd/4"}

# Debajo de este score ni se sugiere como match (puro ruido).
_SCORE_MIN_SUGERENCIA = 55.0
# Arriba de este score se acepta un hint de marca (título de hoja) como match.
_SCORE_MIN_MARCA = 80.0

_SUFIJO_PESO = re.compile(
    r"\s*[-–]?\s*\d+([.,]\d+)?\s*(kgs?|kilos?|grs?|gramos?|ml|lts?|litros?|pzas?)\.?\s*$",
    re.IGNORECASE,
)

_ALIAS_COLUMNAS = {
    "nombre": "nombre",
    "producto": "nombre",
    "descripcion": "nombre",
    "nueva descripcion": "nombre",
    "kg": "contenido_neto",
    "kgs": "contenido_neto",
    "peso": "contenido_neto",
    "contenido": "contenido_neto",
    "presentacion (kg)": "contenido_neto",
    # "Precio" = precio de venta ya definido (ej. Winners) — se usa directo.
    "precio": "precio",
    "precio unitario": "precio",
    "precio de venta": "precio",
    # "Precio de lista"/"Costo" = lo que cuesta comprarlo (ej. formatos de
    # pedido de proveedor) — pasa por el cálculo de margen, no se usa directo.
    "precio de lista": "precio_costo",
    "costo": "precio_costo",
    "precio compra": "precio_costo",
    "precio de compra": "precio_costo",
    "codigo": "codigo_proveedor",
    "codigo del producto": "codigo_proveedor",
    "clave": "codigo_proveedor",
    "cantidad": "cantidad",
    "unidades": "cantidad",
    "unidades (sacos)": "cantidad",
}

_LINEA_COLUMNAS = {
    "nombre_original", "categoria_sugerida", "cantidad", "contenido_neto",
    "precio_costo", "codigo_sat", "codigo_proveedor", "marca_id",
    "producto_match_id", "match_score", "margen_aplicado",
    "precio_venta_sugerido",
}


# --------------------------------------------------------------------------
# Normalización y matching (nombre → producto existente / marca existente)
# --------------------------------------------------------------------------

def _clave_dedup(texto: str) -> str:
    """minúsculas, sin acentos, espacios colapsados — SIN quitar el sufijo de
    peso: dos presentaciones del mismo producto ('Hound Adulto 10 kg' vs
    'Hound Adulto 4 kg') son líneas distintas, no un duplicado."""
    txt = (texto or "").strip().lower()
    txt = "".join(c for c in unicodedata.normalize("NFD", txt) if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", txt).strip()


def _normalizar(texto: str) -> str:
    """Como `_clave_dedup` pero además quita el sufijo de peso — para
    comparar por similitud (fuzzy match), donde sí queremos ignorar el
    tamaño del empaque."""
    return _SUFIJO_PESO.sub("", _clave_dedup(texto)).strip()


def _score(a: str, b: str) -> float:
    return round(difflib.SequenceMatcher(None, _normalizar(a), _normalizar(b)).ratio() * 100, 2)


def buscar_match(nombre: str, candidatos: list[dict]) -> tuple[Optional[int], Optional[float]]:
    """Mejor candidato por nombre (fuzzy) entre productos activos. Solo sugiere."""
    mejor_id, mejor_score = None, 0.0
    for c in candidatos:
        s = _score(nombre, c["nombre"])
        if s > mejor_score:
            mejor_id, mejor_score = c["id"], s
    if mejor_score < _SCORE_MIN_SUGERENCIA:
        return None, (mejor_score or None)
    return mejor_id, mejor_score


def _resolver_marca_por_nombre(hint: Optional[str], marcas: list[dict]) -> Optional[int]:
    if not hint:
        return None
    hint_norm = _normalizar(hint)
    for m in marcas:
        if _normalizar(m["nombre"]) == hint_norm:
            return m["id"]
    mejor_id, mejor_score = None, 0.0
    for m in marcas:
        s = _score(hint, m["nombre"])
        if s > mejor_score:
            mejor_id, mejor_score = m["id"], s
    return mejor_id if mejor_score >= _SCORE_MIN_MARCA else None


def _resolver_categoria_por_nombre(hint: Optional[str], categorias: list[dict]) -> Optional[int]:
    """Igual que `_resolver_marca_por_nombre`, pero contra el catálogo de
    categorías, usando `categoria_sugerida` (texto libre capturado al
    parsear, ej. "Cerdos > Línea Suprema"). Si no hay un match razonable, se
    deja en blanco — no se inventa una categoría (Fase 15, §2.6)."""
    if not hint:
        return None
    hint_norm = _normalizar(hint)
    for c in categorias:
        if _normalizar(c["nombre"]) == hint_norm:
            return c["id"]
    mejor_id, mejor_score = None, 0.0
    for c in categorias:
        s = _score(hint, c["nombre"])
        if s > mejor_score:
            mejor_id, mejor_score = c["id"], s
    return mejor_id if mejor_score >= _SCORE_MIN_MARCA else None


def _precio_sugerido(costo: Optional[float], margen: Optional[float]) -> Optional[float]:
    if costo is None or margen is None:
        return None
    return round(float(costo) * (1 + float(margen) / 100), 2)


async def resolver_margen(marca_id: Optional[int], producto_match_id: Optional[int]) -> float:
    """Override del producto matcheado > margen_default de la marca > 0."""
    if producto_match_id:
        fila = await database.fetch_one(
            select(producto.c.margen_override, producto.c.marca_id).where(producto.c.id == producto_match_id)
        )
        if fila and fila["margen_override"] is not None:
            return float(fila["margen_override"])
        marca_id = marca_id or (fila["marca_id"] if fila else None)
    if marca_id:
        fila = await database.fetch_one(select(marca.c.margen_default).where(marca.c.id == marca_id))
        if fila:
            return float(fila["margen_default"])
    return 0.0


# --------------------------------------------------------------------------
# Parseo de Excel — dos modos, detectados por hoja
# --------------------------------------------------------------------------

def _titulo_hoja(ws) -> Optional[str]:
    """Celda de fuente más grande en la fila 1 → nombre de marca/título de la hoja."""
    mejor_valor, mejor_size = None, 0.0
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if isinstance(cell.value, str) and cell.value.strip():
            size = cell.font.size or 0
            if size > mejor_size:
                mejor_valor, mejor_size = cell.value.strip(), size
    return mejor_valor


def _detectar_header(ws) -> Optional[tuple[int, dict[int, str]]]:
    """Busca una fila con celda 'Nombre'/'Producto'/etc → mapea columnas por alias.
    Se busca en toda la hoja (hasta un límite razonable) porque algunos formatos
    reales traen varias filas de título/metadatos antes del encabezado real
    (ej. un "Formato de Pedidos" con el encabezado hasta la fila 14)."""
    for row in range(1, min(200, ws.max_row + 1)):
        mapeo: dict[int, str] = {}
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if not isinstance(val, str):
                continue
            campo = _ALIAS_COLUMNAS.get(_normalizar(val))
            if campo:
                mapeo[col] = campo
        if "nombre" in mapeo.values():
            return row, mapeo
    return None


def _parsear_hoja_tabla(ws, header_row: int, mapeo: dict[int, str], marca_hint: Optional[str]) -> list[dict]:
    lineas = []
    for row in range(header_row + 1, ws.max_row + 1):
        fila = {campo: ws.cell(row=row, column=col).value for col, campo in mapeo.items()}
        nombre = fila.get("nombre")
        # Filas de categoría dentro de la tabla (ej. "LINEA ESENCIAL", "AVES DE
        # POSTURA") traen texto en OTRA columna pero no en la de nombre — se
        # saltan solas aquí. También se saltan filas sin nombre o de "total".
        if not isinstance(nombre, str) or not nombre.strip() or nombre.strip().lower() == "total":
            continue
        precio = fila.get("precio")
        precio_costo = fila.get("precio_costo")
        contenido = fila.get("contenido_neto")
        cantidad = fila.get("cantidad")
        codigo = fila.get("codigo_proveedor")
        lineas.append({
            "nombre_original": nombre.strip(),
            "precio_venta_sugerido": float(precio) if isinstance(precio, (int, float)) else None,
            "precio_costo": float(precio_costo) if isinstance(precio_costo, (int, float)) else None,
            "contenido_neto": float(contenido) if isinstance(contenido, (int, float)) else None,
            "cantidad": float(cantidad) if isinstance(cantidad, (int, float)) else None,
            "codigo_proveedor": str(codigo).strip() if codigo is not None and str(codigo).strip() else None,
            "marca_hint": marca_hint,
        })
    return lineas


def _parsear_hoja_bloques(ws, marca_hint: Optional[str]) -> list[dict]:
    """Sin encabezados de tabla: bloques de categoría por columna, detectados por
    estilo (bold+tamaño de fuente), no por texto/posición — ver EMPEZAR_AQUI/plan."""
    lineas = []
    categoria_por_col: dict[int, str] = {}
    linea_por_col: dict[int, str] = {}
    for row in range(2, ws.max_row + 1):  # fila 1 = título de la hoja, se omite
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if not isinstance(val, str) or not val.strip():
                continue
            texto = val.strip()
            bold = bool(cell.font.bold)
            size = cell.font.size or 11
            if bold and size >= 14:
                categoria_por_col[col] = texto
                linea_por_col.pop(col, None)
                continue
            if bold:
                linea_por_col[col] = texto
                continue
            categoria = categoria_por_col.get(col)
            sub = linea_por_col.get(col)
            categoria_sugerida = " > ".join(p for p in [categoria, sub] if p) or None
            lineas.append({
                "nombre_original": texto,
                "categoria_sugerida": categoria_sugerida,
                "marca_hint": marca_hint,
            })
    return lineas


def parsear_excel(contenido: bytes) -> list[dict]:
    """Devuelve líneas crudas (sin tocar la BD). Dedup por nombre normalizado
    dentro del archivo (una hoja de pedido puede repetir la misma lista de precio)."""
    wb = openpyxl.load_workbook(BytesIO(contenido), data_only=True)
    vistos: set[str] = set()
    lineas: list[dict] = []
    for ws in wb.worksheets:
        marca_hint = _titulo_hoja(ws)
        header = _detectar_header(ws)
        crudas = (
            _parsear_hoja_tabla(ws, *header, marca_hint)
            if header else _parsear_hoja_bloques(ws, marca_hint)
        )
        for l in crudas:
            clave = _clave_dedup(l["nombre_original"])
            if not clave or clave in vistos:
                continue
            vistos.add(clave)
            lineas.append(l)
    return lineas


# --------------------------------------------------------------------------
# Parseo de PDF (listas de precios de proveedor/distribuidor)
# --------------------------------------------------------------------------

def _a_float_pdf(valor) -> Optional[float]:
    """Convierte celdas numéricas de tabla de PDF ('13,812.00', '8 ,940.00'
    con espacio suelto antes de la coma) a float."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    limpio = str(valor).strip().replace(" ", "").replace(",", "")
    if not limpio:
        return None
    try:
        return float(limpio)
    except ValueError:
        return None


def _clasificar_columna_pdf(texto: str) -> Optional[str]:
    """Clasifica una columna de tabla de PDF por su encabezado (ya combinado
    y normalizado). No usa un diccionario de alias exacto como el de Excel
    porque los encabezados de PDF suelen venir partidos en 2+ filas con texto
    que se concatena de forma impredecible (ej. 'Precio' + 'Tonelada')."""
    if "descripcion" in texto:
        return "nombre"
    if "item" in texto:
        return "codigo_proveedor"
    if "tonelada" in texto:
        return None  # precio derivado/redundante (se puede recalcular), no se usa
    if "saco" in texto and "kg" in texto:
        return "contenido_neto"
    if "saco" in texto or "bulto" in texto:
        return "precio_costo"
    if "kg" in texto or "peso" in texto:
        return "contenido_neto"
    if "precio" in texto or "costo" in texto:
        return "precio_costo"
    return None


def _combinar_encabezado_pdf(tabla: list[list], filas_header: int = 2) -> dict[int, str]:
    n_cols = max((len(f) for f in tabla[:filas_header]), default=0)
    mapeo: dict[int, str] = {}
    for c in range(n_cols):
        partes = [str(f[c]) for f in tabla[:filas_header] if c < len(f) and f[c]]
        campo = _clasificar_columna_pdf(_clave_dedup(" ".join(partes)))
        if campo:
            mapeo[c] = campo
    return mapeo


def parsear_pdf(contenido: bytes) -> list[dict]:
    """Extrae líneas de tablas de PDF (listas de precio de proveedor/distribuidor).
    Siempre trata el precio encontrado como COSTO (nunca precio de venta directo)
    — estas listas son del proveedor hacia el negocio, no de venta al público."""
    import pdfplumber

    lineas: list[dict] = []
    vistos: set[str] = set()
    categoria_actual: Optional[str] = None
    # El mapeo de columnas se mantiene como estado ENTRE fragmentos de tabla:
    # pdfplumber parte una misma página en varias "tablas" (una por sección/
    # categoría, ej. POLLORINA, CAPORINA...) y solo la primera trae el
    # encabezado repetido — las siguientes son continuaciones de puros datos.
    # Tratar las primeras 2 filas de CADA fragmento como encabezado (como se
    # hacía antes) descartaba productos reales y perdía secciones enteras.
    mapeo_activo: dict[int, str] = {}

    with pdfplumber.open(BytesIO(contenido)) as pdf:
        for page in pdf.pages:
            for tabla in page.extract_tables():
                if not tabla:
                    continue
                candidato = _combinar_encabezado_pdf(tabla, 2)
                if "nombre" in candidato.values():
                    mapeo_activo = candidato
                    filas_datos = tabla[2:]
                else:
                    filas_datos = tabla  # continuación: son puros datos, no encabezado

                if "nombre" not in mapeo_activo.values():
                    continue  # todavía no se detecta ningún encabezado válido

                for fila in filas_datos:
                    valores = {campo: (fila[c] if c < len(fila) else None) for c, campo in mapeo_activo.items()}
                    nombre = valores.get("nombre")
                    if not nombre or not str(nombre).strip():
                        # Fila de categoría (ej. "POLLORINA") — el texto viene en
                        # cualquier otra columna de ese renglón. Se guarda como
                        # categoría vigente para las líneas siguientes.
                        texto_libre = next((str(v).strip() for v in fila if v and str(v).strip()), None)
                        if texto_libre:
                            categoria_actual = texto_libre
                        continue

                    clave = _clave_dedup(str(nombre))
                    if not clave or clave in vistos:
                        continue
                    vistos.add(clave)

                    codigo = valores.get("codigo_proveedor")
                    lineas.append({
                        "nombre_original": str(nombre).strip(),
                        "categoria_sugerida": categoria_actual,
                        "codigo_proveedor": str(codigo).strip() if codigo and str(codigo).strip() else None,
                        "contenido_neto": _a_float_pdf(valores.get("contenido_neto")),
                        "precio_costo": _a_float_pdf(valores.get("precio_costo")),
                        "marca_hint": None,
                    })
    return lineas


# --------------------------------------------------------------------------
# Parseo de XML (CFDI 4.0 — factura de compra)
# --------------------------------------------------------------------------

def parsear_xml_cfdi(contenido: bytes) -> dict:
    root = ET.fromstring(contenido)
    emisor = root.find("cfdi:Emisor", CFDI_NS)
    proveedor = emisor.get("Nombre") if emisor is not None else None

    lineas = []
    conceptos = root.find("cfdi:Conceptos", CFDI_NS)
    for concepto in (conceptos.findall("cfdi:Concepto", CFDI_NS) if conceptos is not None else []):
        cantidad = float(concepto.get("Cantidad") or 0)
        importe = float(concepto.get("Importe") or 0)
        descuento = float(concepto.get("Descuento") or 0)
        precio_costo = round((importe - descuento) / cantidad, 4) if cantidad else None
        lineas.append({
            "nombre_original": (concepto.get("Descripcion") or "").strip(),
            "cantidad": cantidad,
            "precio_costo": precio_costo,
            "codigo_sat": concepto.get("ClaveProdServ"),
            "codigo_proveedor": concepto.get("NoIdentificacion"),
        })
    return {"proveedor": proveedor, "lineas": lineas}


# --------------------------------------------------------------------------
# Enriquecimiento (matching + margen) — ya con acceso a BD
# --------------------------------------------------------------------------

async def _productos_activos() -> list[dict]:
    filas = await database.fetch_all(
        select(producto.c.id, producto.c.nombre).where(producto.c.activo == True)
    )
    return [dict(f) for f in filas]


async def _marcas_existentes() -> list[dict]:
    filas = await database.fetch_all(select(marca.c.id, marca.c.nombre))
    return [dict(f) for f in filas]


async def _categorias_existentes() -> list[dict]:
    filas = await database.fetch_all(select(categoria.c.id, categoria.c.nombre))
    return [dict(f) for f in filas]


async def _sku_disponible(codigo: Optional[str], excluir_producto_id: Optional[int] = None) -> bool:
    """`producto.sku` es único — antes de usar el código de proveedor de una
    línea de importación como sku, hay que confirmar que ningún OTRO producto
    ya lo tenga (dos proveedores distintos pueden coincidir en un código)."""
    if not codigo:
        return False
    query = select(producto.c.id).where(producto.c.sku == codigo)
    if excluir_producto_id is not None:
        query = query.where(producto.c.id != excluir_producto_id)
    return await database.fetch_one(query) is None


async def enriquecer_lineas(lineas_crudas: list[dict], marca_default_id: Optional[int] = None) -> list[dict]:
    marcas = await _marcas_existentes()
    productos = await _productos_activos()

    resultado = []
    for cruda in lineas_crudas:
        cruda = dict(cruda)
        marca_hint = cruda.pop("marca_hint", None)
        marca_id = _resolver_marca_por_nombre(marca_hint, marcas) or marca_default_id

        producto_match_id, score = buscar_match(cruda["nombre_original"], productos)

        precio_venta_sugerido = cruda.get("precio_venta_sugerido")  # Excel-tabla: ya viene
        margen_aplicado = None
        if precio_venta_sugerido is None and cruda.get("precio_costo") is not None:
            margen_aplicado = await resolver_margen(marca_id, producto_match_id)
            precio_venta_sugerido = _precio_sugerido(cruda["precio_costo"], margen_aplicado)

        cruda["marca_id"] = marca_id
        cruda["producto_match_id"] = producto_match_id
        cruda["match_score"] = score
        cruda["margen_aplicado"] = margen_aplicado
        cruda["precio_venta_sugerido"] = precio_venta_sugerido
        resultado.append(cruda)
    return resultado


def _solo_columnas_linea(l: dict) -> dict:
    return {k: v for k, v in l.items() if k in _LINEA_COLUMNAS}


# --------------------------------------------------------------------------
# Alta de lotes
# --------------------------------------------------------------------------

async def crear_lote_excel(
    contenido: bytes, nombre_archivo: str, usuario_id: int, marca_default_id: Optional[int] = None
) -> int:
    crudas = parsear_excel(contenido)
    enriquecidas = await enriquecer_lineas(crudas, marca_default_id=marca_default_id)
    async with database.transaction():
        lote_id = await database.execute(
            importacion_lote.insert().values(
                tipo=IMPORT_TIPO_EXCEL, nombre_archivo=nombre_archivo, usuario_id=usuario_id,
                marca_default_id=marca_default_id,
            )
        )
        for l in enriquecidas:
            await database.execute(
                importacion_linea.insert().values(lote_id=lote_id, **_solo_columnas_linea(l))
            )
    return lote_id


async def crear_lote_xml(
    contenido: bytes, nombre_archivo: str, usuario_id: int, marca_default_id: Optional[int] = None
) -> int:
    parsed = parsear_xml_cfdi(contenido)
    enriquecidas = await enriquecer_lineas(parsed["lineas"], marca_default_id=marca_default_id)
    async with database.transaction():
        lote_id = await database.execute(
            importacion_lote.insert().values(
                tipo=IMPORT_TIPO_XML, nombre_archivo=nombre_archivo,
                proveedor=parsed["proveedor"], usuario_id=usuario_id,
                marca_default_id=marca_default_id,
            )
        )
        for l in enriquecidas:
            await database.execute(
                importacion_linea.insert().values(lote_id=lote_id, **_solo_columnas_linea(l))
            )
    return lote_id


async def crear_lote_pdf(
    contenido: bytes, nombre_archivo: str, usuario_id: int, marca_default_id: Optional[int] = None
) -> int:
    crudas = parsear_pdf(contenido)
    enriquecidas = await enriquecer_lineas(crudas, marca_default_id=marca_default_id)
    async with database.transaction():
        lote_id = await database.execute(
            importacion_lote.insert().values(
                tipo=IMPORT_TIPO_PDF, nombre_archivo=nombre_archivo, usuario_id=usuario_id,
                marca_default_id=marca_default_id,
            )
        )
        for l in enriquecidas:
            await database.execute(
                importacion_linea.insert().values(lote_id=lote_id, **_solo_columnas_linea(l))
            )
    return lote_id


# --------------------------------------------------------------------------
# Edición de línea (vista previa)
# --------------------------------------------------------------------------

async def actualizar_linea(lote_id: int, linea_id: int, cambios: dict) -> None:
    linea = await database.fetch_one(
        importacion_linea.select().where(
            (importacion_linea.c.id == linea_id) & (importacion_linea.c.lote_id == lote_id)
        )
    )
    if not linea:
        raise HTTPException(404, "Línea no encontrada")
    if not cambios:
        return
    if "decision" in cambios and cambios["decision"] not in IMPORT_DECISIONES:
        raise HTTPException(400, "Decisión inválida")

    datos = dict(cambios)
    costo = float(linea["precio_costo"]) if linea["precio_costo"] is not None else None

    # Si cambia la marca (y no se manda margen/precio explícito), resolver el
    # margen de esa marca y recalcular el precio sugerido — es el punto central
    # del diseño: asignar marca a una línea de factura debe aplicar su margen.
    if "marca_id" in datos and "margen_aplicado" not in datos and "precio_venta_sugerido" not in datos and costo is not None:
        datos["margen_aplicado"] = await resolver_margen(datos["marca_id"], linea["producto_match_id"])
        datos["precio_venta_sugerido"] = _precio_sugerido(costo, datos["margen_aplicado"])

    # Si se ajusta el margen sin mandar precio explícito, recalcular el precio sugerido.
    if "margen_aplicado" in datos and "precio_venta_sugerido" not in datos:
        datos["precio_venta_sugerido"] = _precio_sugerido(costo, datos["margen_aplicado"])

    await database.execute(
        importacion_linea.update().where(importacion_linea.c.id == linea_id).values(**datos)
    )


async def eliminar_linea(lote_id: int, linea_id: int) -> None:
    """Quita una línea del lote por completo (no solo 'ignorar') — para errores
    de parseo (ej. texto que no era un producto real) que el dueño quiere
    limpiar de la vista previa antes de confirmar."""
    resultado = await database.execute(
        importacion_linea.delete().where(
            (importacion_linea.c.id == linea_id) & (importacion_linea.c.lote_id == lote_id)
        )
    )
    if resultado == 0:
        raise HTTPException(404, "Línea no encontrada")


async def cancelar_lote(lote_id: int) -> None:
    resultado = await database.execute(
        importacion_lote.update()
        .where((importacion_lote.c.id == lote_id) & (importacion_lote.c.estado == IMPORT_ESTADO_REVISION))
        .values(estado=IMPORT_ESTADO_CANCELADO)
    )
    if resultado == 0:
        raise HTTPException(404, "Lote no encontrado o ya no está en revisión")


# --------------------------------------------------------------------------
# Confirmación — única función que escribe en el catálogo/inventario real
# --------------------------------------------------------------------------

async def confirmar_lote(lote_id: int, usuario: dict, generar_ingreso: bool, sucursal_id: Optional[int]) -> dict:
    lote = await database.fetch_one(importacion_lote.select().where(importacion_lote.c.id == lote_id))
    if not lote:
        raise HTTPException(404, "Lote no encontrado")
    if lote["estado"] != IMPORT_ESTADO_REVISION:
        raise HTTPException(409, "Este lote ya fue confirmado o cancelado")
    if generar_ingreso and not sucursal_id:
        raise HTTPException(400, "Debes elegir una sucursal para generar el ingreso de inventario")

    lineas = await database.fetch_all(
        importacion_linea.select().where(importacion_linea.c.lote_id == lote_id)
    )
    categorias = await _categorias_existentes()

    creados = vinculados = ignorados = ingresos = 0

    async with database.transaction():
        for linea in lineas:
            if linea["decision"] in (IMPORT_DECISION_IGNORAR, "pendiente"):
                ignorados += 1
                continue

            producto_id = None

            if linea["decision"] == IMPORT_DECISION_CREAR:
                tiene_precio = linea["precio_venta_sugerido"] is not None
                contenido = float(linea["contenido_neto"]) if linea["contenido_neto"] is not None else 1.0
                unidad = "Bulto" if linea["contenido_neto"] is not None else "pza"
                # Código de proveedor (NoIdentificacion del CFDI, Item del PDF,
                # Código del Excel) como sku, si no está ya usado por otro producto.
                sku = linea["codigo_proveedor"] if await _sku_disponible(linea["codigo_proveedor"]) else None
                categoria_id = _resolver_categoria_por_nombre(linea["categoria_sugerida"], categorias)
                producto_id = await database.execute(
                    producto.insert().values(
                        nombre=linea["nombre_original"],
                        sku=sku,
                        marca_id=linea["marca_id"],
                        categoria_id=categoria_id,
                        unidad_medida=unidad,
                        contenido_neto=contenido,
                        se_vende_a_granel=False,
                        precio_base=float(linea["precio_venta_sugerido"]) if tiene_precio else 0.0,
                        costo=float(linea["precio_costo"]) if linea["precio_costo"] is not None else None,
                        # Sin precio real (Excel de solo-nombre): se crea suspendido para
                        # que no sea vendible a $0 hasta que el dueño lo precie a mano.
                        activo=tiene_precio,
                        # Explícito: el `default=5.0` de SQLAlchemy Core no siempre se
                        # aplica en este camino de inserción (visto en pruebas — quedaba
                        # NULL y rompía la lectura del producto contra el schema).
                        stock_minimo=5.0,
                    )
                )
                await database.execute(
                    importacion_linea.update().where(importacion_linea.c.id == linea["id"])
                    .values(producto_creado_id=producto_id)
                )
                creados += 1

            elif linea["decision"] == IMPORT_DECISION_VINCULAR:
                if not linea["producto_match_id"]:
                    raise HTTPException(
                        400, f"La línea {linea['id']} está marcada 'vincular' pero no tiene producto asignado"
                    )
                producto_id = linea["producto_match_id"]
                producto_previo = await database.fetch_one(
                    select(producto).where(producto.c.id == producto_id)
                )
                actualizacion = {}
                if linea["precio_costo"] is not None:
                    actualizacion["costo"] = float(linea["precio_costo"])
                if linea["actualizar_precio"] and linea["precio_venta_sugerido"] is not None:
                    actualizacion["precio_base"] = float(linea["precio_venta_sugerido"])
                # Solo se rellena el sku si el producto todavía no tiene uno —
                # nunca se pisa un sku ya capturado a mano o de una importación previa.
                if (
                    producto_previo and not producto_previo["sku"]
                    and await _sku_disponible(linea["codigo_proveedor"], excluir_producto_id=producto_id)
                ):
                    actualizacion["sku"] = linea["codigo_proveedor"]
                if actualizacion:
                    await database.execute(
                        producto.update().where(producto.c.id == producto_id).values(**actualizacion)
                    )
                    await registrar_cambio_precio(
                        producto_id=producto_id,
                        usuario_id=usuario["id"],
                        costo_anterior=producto_previo["costo"],
                        costo_nuevo=actualizacion.get("costo"),
                        precio_anterior=producto_previo["precio_base"],
                        precio_nuevo=actualizacion.get("precio_base"),
                        origen="importacion",
                        lote_id=lote_id,
                    )
                vinculados += 1

            if generar_ingreso and producto_id and linea["cantidad"]:
                prod_row = await database.fetch_one(select(producto).where(producto.c.id == producto_id))
                contenido_neto = float(prod_row["contenido_neto"] or 1)
                kilos_reales = float(linea["cantidad"]) * contenido_neto

                inv_actual = await database.fetch_one(
                    select(inventario).where(
                        (inventario.c.producto_id == producto_id) & (inventario.c.sucursal_id == sucursal_id)
                    )
                )
                cantidad_anterior = float(inv_actual["cantidad"]) if inv_actual else 0.0
                nueva_cantidad = cantidad_anterior + kilos_reales

                if inv_actual:
                    await database.execute(
                        inventario.update().where(inventario.c.id == inv_actual["id"])
                        .values(cantidad=nueva_cantidad, fecha_actualizacion=datetime.now(timezone.utc))
                    )
                else:
                    await database.execute(
                        inventario.insert().values(
                            producto_id=producto_id, sucursal_id=sucursal_id,
                            cantidad=kilos_reales, fecha_actualizacion=datetime.now(timezone.utc),
                        )
                    )

                await database.execute(
                    ingreso_inventario.insert().values(
                        producto_id=producto_id, sucursal_id=sucursal_id,
                        cantidad=linea["cantidad"], usuario_id=usuario["id"],
                    )
                )
                await registrar_movimiento(
                    sucursal_id=sucursal_id, usuario_id=usuario["id"], producto_id=producto_id,
                    tipo_movimiento=MOV_COMPRA,
                    cantidad_anterior=cantidad_anterior, cantidad_movida=kilos_reales,
                    cantidad_nueva=nueva_cantidad,
                    motivo=f"Importación factura ({lote['nombre_archivo']})",
                )
                ingresos += 1

        await database.execute(
            importacion_lote.update().where(importacion_lote.c.id == lote_id).values(
                estado=IMPORT_ESTADO_CONFIRMADO,
                fecha_confirmacion=datetime.now(timezone.utc),
                generar_ingreso=generar_ingreso,
                sucursal_id=sucursal_id,
            )
        )

    return {
        "productos_creados": creados,
        "productos_vinculados": vinculados,
        "lineas_ignoradas": ignorados,
        "ingresos_generados": ingresos,
    }
