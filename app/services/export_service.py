"""Generación de reportes en Excel (openpyxl) para descarga desde el panel."""
from collections import defaultdict
from io import BytesIO
from typing import Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def generar_excel(encabezados: Sequence[str], filas: Sequence[Sequence], nombre_hoja: str = "Reporte") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja[:31]  # límite de Excel para nombres de hoja

    ws.append(list(encabezados))
    for celda in ws[1]:
        celda.font = Font(bold=True)

    for fila in filas:
        ws.append(list(fila))

    for columna in ws.columns:
        largo = max((len(str(c.value)) for c in columna if c.value is not None), default=8)
        ws.column_dimensions[columna[0].column_letter].width = min(largo + 2, 45)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _aclarar(rgb_hex: str, tint: float) -> str:
    """Aproxima el "tinte" de Excel (mezcla hacia blanco) — no es idéntico al
    algoritmo HSL exacto de Office, pero visualmente equivalente."""
    r, g, b = int(rgb_hex[0:2], 16), int(rgb_hex[2:4], 16), int(rgb_hex[4:6], 16)
    r = round(r + (255 - r) * tint)
    g = round(g + (255 - g) * tint)
    b = round(b + (255 - b) * tint)
    return f"{r:02X}{g:02X}{b:02X}"


_COLUMNAS_PANEL_EXPORT = {"izq": (1, 3, 4), "der": (5, 7, 8)}  # (col_nombre, col_precio, col_sep)


_CENTRO = Alignment(horizontal="center", vertical="center")

# Márgenes reducidos tomados celda-por-celda de los .xlsx originales del
# dueño (openpyxl, en pulgadas) — así el Excel exportado imprime igual de
# compacto que el que él ya usa, sin páginas extra en blanco.
_MARGENES_IMPRESION = {"left": 0.24, "right": 0.2, "top": 0.28, "bottom": 0.2, "header": 0.31, "footer": 0.31}


def generar_excel_plantilla(marca_nombre: str, color_hex: str, filas: Sequence[dict]) -> bytes:
    """Reconstruye el layout "menú a dos paneles" de Lista Agromas/Api-Aba
    (Fase 16, PLAN_FASE16.md §4.5) a partir de `lista_plantilla_fila` — no
    copia el .xlsx original celda por celda porque la estructura es editable
    desde la web (§6.6) y puede ya no coincidir con las filas del archivo
    fuente; en cambio reproduce el mismo esquema visual (título centrado,
    encabezados centrados con tinte de marca, nombres de producto centrados,
    paneles A-D/E-H, bordes de celda) con los datos vigentes — replicando lo
    que el dueño imprime hoy: texto centrado y área de impresión ajustada
    (sin filas en blanco de sobra) con márgenes reducidos.

    `filas`: dicts con hoja, panel, orden, tipo, nivel, texto, precio_base.
    """
    color_rgb = color_hex.lstrip("#").upper()
    borde = Border(*(Side(style="thin"),) * 4)

    por_hoja: dict[str, list[dict]] = defaultdict(list)
    for f in filas:
        por_hoja[f["hoja"]].append(f)

    wb = Workbook()
    wb.remove(wb.active)

    for nombre_hoja, filas_hoja in por_hoja.items():
        ws = wb.create_sheet(nombre_hoja[:31])
        for col, ancho in zip("ABCDEFGH", (15, 15, 11, 3, 15, 15, 11, 3)):
            ws.column_dimensions[col].width = ancho

        ws.merge_cells("A1:H2")
        titulo = ws.cell(row=1, column=1, value=marca_nombre)
        titulo.font = Font(size=26, bold=True)
        titulo.fill = PatternFill("solid", fgColor=f"FF{color_rgb}")
        titulo.alignment = _CENTRO
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 15

        por_panel: dict[str, list[dict]] = defaultdict(list)
        for f in filas_hoja:
            por_panel[f["panel"]].append(f)

        ultima_fila = 2
        for panel, filas_panel in por_panel.items():
            col_nombre, col_precio, col_sep = _COLUMNAS_PANEL_EXPORT[panel]
            fila_excel = 3
            for f in sorted(filas_panel, key=lambda x: x["orden"]):
                if f["tipo"] == "encabezado":
                    nivel = f["nivel"] or 2
                    tint = {1: 0.0, 2: 0.4}.get(nivel, 0.8)
                    ws.merge_cells(start_row=fila_excel, start_column=col_nombre, end_row=fila_excel, end_column=col_sep)
                    celda = ws.cell(row=fila_excel, column=col_nombre, value=f["texto"])
                    celda.font = Font(size=16 if nivel <= 2 else 11, bold=True)
                    celda.fill = PatternFill("solid", fgColor=f"FF{_aclarar(color_rgb, tint)}")
                    celda.alignment = _CENTRO
                    ws.row_dimensions[fila_excel].height = 20.1 if nivel <= 2 else 15
                else:
                    ws.merge_cells(start_row=fila_excel, start_column=col_nombre, end_row=fila_excel, end_column=col_nombre + 1)
                    celda_nombre = ws.cell(row=fila_excel, column=col_nombre, value=f["texto"])
                    celda_nombre.font = Font(size=11)
                    celda_nombre.alignment = _CENTRO
                    celda_nombre.border = borde
                    celda_sep = ws.cell(row=fila_excel, column=col_nombre + 1)
                    celda_sep.border = borde
                    precio: Optional[float] = f.get("precio_base")
                    celda_precio = ws.cell(row=fila_excel, column=col_precio, value=float(precio) if precio is not None else None)
                    celda_precio.border = borde
                    celda_precio.alignment = _CENTRO
                    if precio is not None:
                        celda_precio.number_format = '"$"#,##0.00'
                    ws.row_dimensions[fila_excel].height = 15
                fila_excel += 1
            ultima_fila = max(ultima_fila, fila_excel - 1)

        # Área de impresión ajustada al contenido real (sin páginas extra en
        # blanco) + márgenes reducidos, igual que el dueño ya imprime hoy.
        ws.print_area = f"A1:H{ultima_fila}"
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        pm = ws.page_margins
        for lado, valor in _MARGENES_IMPRESION.items():
            setattr(pm, lado, valor)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
