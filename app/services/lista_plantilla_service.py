"""Extrae la estructura de las plantillas Excel (Lista Agromas / Lista Api-Aba)
a una lista de nodos ordenados — Fase 16, ver PLAN_FASE16.md §4.1.

Cada hoja del Excel es un "catálogo tipo menú a dos columnas": columna A/C
(nombre/precio, panel izquierdo) y columna E/G (panel derecho). Dentro de
cada panel, las filas forman un esquema (outline) de profundidad variable:
encabezados en negrita con relleno de color de marca (nivel 1 = tinte 0,
nivel 2 = tinte medio, nivel 3 = tinte claro) seguidos de filas de producto
sin relleno. No hay correspondencia fila-a-fila entre panel izquierdo y
derecho — son dos listas independientes que comparten la rejilla de Excel
solo para acomodarse en una página impresa.
"""
import re
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

PANEL_IZQ = "izq"
PANEL_DER = "der"
_COLUMNAS_PANEL = {PANEL_IZQ: (1, 3), PANEL_DER: (5, 7)}  # (col_nombre, col_precio)

TIPO_ENCABEZADO = "encabezado"
TIPO_PRODUCTO = "producto"


@dataclass
class NodoPlantilla:
    hoja: str
    panel: str
    orden: int
    tipo: str
    nivel: Optional[int]
    texto: str


def _rango_fusionado(ws: Worksheet, row: int, col: int):
    for rango in ws.merged_cells.ranges:
        if rango.min_row <= row <= rango.max_row and rango.min_col <= col <= rango.max_col:
            return rango
    return None


def _nivel_encabezado(cell) -> Optional[int]:
    """None si la celda es una fila de producto normal (sin relleno de marca).
    1/2/3 según qué tan oscuro es el tinte del relleno (1 = color pleno)."""
    if not cell.font or not cell.font.bold:
        return None
    fill = cell.fill
    if not fill or not fill.fgColor or fill.fgColor.type != "theme":
        return None
    tint = fill.fgColor.tint or 0
    if abs(tint) < 1e-6:
        return 1
    if tint <= 0.5:
        return 2
    return 3


def _extraer_panel(ws: Worksheet, hoja: str, panel: str) -> list[NodoPlantilla]:
    col_nombre, _ = _COLUMNAS_PANEL[panel]
    nodos: list[NodoPlantilla] = []
    orden = 0
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_nombre)
        valor = cell.value
        if valor is None or (isinstance(valor, str) and not valor.strip()):
            continue

        rango = _rango_fusionado(ws, row, col_nombre)
        if rango and rango.min_col <= 4 and rango.max_col >= 5:
            # Título de hoja completa (cruza los dos paneles) — no es un
            # renglón de este panel, es metadata de la hoja completa.
            continue

        orden += 1
        nivel = _nivel_encabezado(cell)
        texto = str(valor).strip()
        if nivel is not None:
            nodos.append(NodoPlantilla(hoja, panel, orden, TIPO_ENCABEZADO, nivel, texto))
        else:
            nodos.append(NodoPlantilla(hoja, panel, orden, TIPO_PRODUCTO, None, texto))
    return nodos


def parsear_plantilla(ruta_xlsx: str) -> list[NodoPlantilla]:
    """Devuelve todos los nodos (de todas las hojas) de un archivo de plantilla,
    en el orden en que deben insertarse."""
    wb = load_workbook(ruta_xlsx, data_only=True)
    nodos: list[NodoPlantilla] = []
    for hoja in wb.sheetnames:
        ws = wb[hoja]
        nodos += _extraer_panel(ws, hoja, PANEL_IZQ)
        nodos += _extraer_panel(ws, hoja, PANEL_DER)
    return nodos


# ---------------------------------------------------------------------------
# Vínculo fila de plantilla ↔ producto del catálogo (PLAN_FASE16.md §4.2):
# el dueño decidió alinear nombres en el origen en vez de un matching difuso
# "adivinado" en runtime, así que esto solo resuelve por coincidencia de
# palabras normalizadas (no similitud aproximada) y deja sin vincular
# (precio vacío) cualquier caso ambiguo — no "adivina" entre varios
# candidatos igual de válidos.
# ---------------------------------------------------------------------------
_UNIDADES = {"KG", "KGS", "GR", "G", "ML", "LT", "L", "PZA"}
_VACIAS = {"DE", "LA", "EL", "LOS", "LAS", "Y", "CON", "A", "PARA"}


def _tokens(texto: str) -> frozenset[str]:
    t = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode()
    t = t.upper()
    # Separa dígitos pegados a letras (ej. "25KG", "5KGS") antes de tokenizar,
    # si no "KG"/"KGS" nunca se reconoce como unidad a quitar.
    t = re.sub(r"(\d)([A-Z])", r"\1 \2", t)
    t = re.sub(r"([A-Z])(\d)", r"\1 \2", t)
    t = re.sub(r"[^A-Z0-9]+", " ", t)  # también tira el "%" — 20% y 20 deben matchear igual
    return frozenset(tok for tok in t.split() if tok not in _UNIDADES and tok not in _VACIAS)


def _es_digito(tok: str) -> bool:
    return tok.isdigit()


def _cubierta(palabra: str, permitidas: frozenset[str]) -> bool:
    """Compara con tolerancia a singular/plural español (CERDO~CERDOS)."""
    if palabra in permitidas:
        return True
    return any(
        len(palabra) >= 3 and len(p) >= 3 and abs(len(palabra) - len(p)) <= 1
        and (palabra.startswith(p) or p.startswith(palabra))
        for p in permitidas
    )


def _construir_ancestros(nodos: list[NodoPlantilla]) -> list[list[str]]:
    """Para cada posición, la lista de textos de encabezado (nivel 1..N) bajo
    los que cae ese nodo dentro de su propio hoja+panel."""
    resultado: list[list[str]] = [[] for _ in nodos]
    pila: dict[int, str] = {}
    grupo_actual = None
    for i, n in enumerate(nodos):
        clave = (n.hoja, n.panel)
        if clave != grupo_actual:
            pila = {}
            grupo_actual = clave
        if n.tipo == TIPO_ENCABEZADO:
            pila = {k: v for k, v in pila.items() if k < n.nivel}
            pila[n.nivel] = n.texto
        else:
            resultado[i] = [pila[k] for k in sorted(pila)]
    return resultado


def resolver_vinculos(nodos: list[NodoPlantilla], productos: list[dict]) -> list[Optional[int]]:
    """productos: [{"id":.., "nombre":.., "sku":..}, ...] de UNA marca.
    Devuelve una lista paralela a `nodos`: producto_id o None por posición
    (None también para los nodos tipo 'encabezado').

    Regla (deliberadamente conservadora — nunca "adivina" entre productos de
    otra especie/línea solo porque comparten una palabra genérica como
    "Crecimiento" o "Engorda"):
    1. Todas las palabras del texto de plantilla deben estar en el nombre del
       producto (ej. "Crecimiento" ⊆ "CERDO SUPREMA CRECIMIENTO 25KG").
    2. Cualquier palabra EXTRA que traiga el producto más allá de esas debe
       explicarse por el contexto (los encabezados sección/línea bajo los que
       cae la fila en la plantilla, ej. "Cerdos"/"Suprema") o ser un tamaño
       (dígito). Si el producto trae una palabra extra que no está en el
       contexto (ej. "Pollas" al intentar matchear el "Desarrollo" de
       Cerdos), NO es candidato — evita cruces entre especies.
    3. Si la plantilla no trae tamaño explícito, no se matchea contra una
       variante "chica" (≤10, ej. 5kg) — esas quedan para su propio renglón.
    4. Si más de un producto sigue calificando, se deja SIN vincular (precio
       vacío) en vez de adivinar.
    """
    ancestros = _construir_ancestros(nodos)
    conteo_leaf: dict[frozenset, int] = defaultdict(int)
    for n in nodos:
        if n.tipo == TIPO_PRODUCTO:
            conteo_leaf[_tokens(n.texto)] += 1

    prods = [{"id": p["id"], "sku": p.get("sku"), "palabras": _tokens(p["nombre"])} for p in productos]
    usados: set[int] = set()
    resultado: list[Optional[int]] = [None] * len(nodos)

    for i, n in enumerate(nodos):
        if n.tipo != TIPO_PRODUCTO:
            continue
        palabras_leaf = _tokens(n.texto)
        if not palabras_leaf:
            continue
        digitos_leaf = {t for t in palabras_leaf if _es_digito(t)}

        palabras_contexto: frozenset[str] = frozenset()
        for texto_ancestro in ancestros[i]:
            palabras_contexto |= _tokens(texto_ancestro)
        permitidas = palabras_leaf | palabras_contexto

        candidatos = []
        for p in prods:
            if p["id"] in usados or not (palabras_leaf <= p["palabras"]):
                continue
            extra = p["palabras"] - palabras_leaf
            extra_digitos = {t for t in extra if _es_digito(t)}
            extra_no_digito = extra - extra_digitos

            if not all(_cubierta(palabra, permitidas) for palabra in extra_no_digito):
                continue  # palabra extra sin explicación de contexto -> no es candidato

            if digitos_leaf:
                if not (digitos_leaf <= extra_digitos or digitos_leaf <= p["palabras"]):
                    continue
            else:
                # Plantilla sin tamaño explícito: no matchear contra una
                # variante "chica" (ej. 5kg) que sí trae un tamaño propio —
                # esas quedan reservadas para su propio renglón con dígito.
                if any(int(d) <= 10 for d in extra_digitos):
                    continue
            candidatos.append(p)

        if len(candidatos) > 1:
            # Empate entre duplicados EXACTOS (mismas palabras, ej. dos
            # productos con nombre idéntico creados dos veces por error):
            # preferir el que ya tiene sku capturado (más completo) en vez
            # de dejarlo sin vincular.
            mismas_palabras = {frozenset(c["palabras"]) for c in candidatos}
            if len(mismas_palabras) == 1:
                con_sku = [c for c in candidatos if c["sku"]]
                if len(con_sku) == 1:
                    candidatos = con_sku

        if len(candidatos) == 1:
            resultado[i] = candidatos[0]["id"]
            usados.add(candidatos[0]["id"])
        # 0 candidatos -> sin precio (no existe en catálogo confirmado).
        # >1 candidatos -> ambiguo, se deja sin vincular a propósito.

    return resultado
