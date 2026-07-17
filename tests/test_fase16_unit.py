"""
Pruebas unitarias (sin BD) de Fase 16 — módulo Listas gráfico.

Cubre lo que sí es lógica pura: el parser de estructura (`parsear_plantilla`,
contra un .xlsx sintético armado en el propio test, no contra los archivos
reales de `ejemplos_importacion/`) y el resolver de vínculos por nombre
(`resolver_vinculos`). El resto (export Excel con estilos, endpoints con BD)
sigue verificándose manualmente contra la BD real, como documenta
EMPEZAR_AQUI.md.

Ejecutar:  pytest -q
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.lista_plantilla_service import (
    NodoPlantilla, TIPO_ENCABEZADO, TIPO_PRODUCTO,
    parsear_plantilla, resolver_vinculos,
)


# --- parsear_plantilla: construir un .xlsx sintético de dos paneles ---

def _fill(theme: int, tint: float):
    from openpyxl.styles.colors import Color
    return PatternFill("solid", fgColor=Color(theme=theme, tint=tint, type="theme"))


def _escribir_plantilla_sintetica(ruta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"

    # Título de hoja completa (cruza los 2 paneles) -> debe IGNORARSE.
    ws.merge_cells("A1:H2")
    ws["A1"] = "MarcaTest"
    ws["A1"].font = Font(bold=True, size=26)
    ws["A1"].fill = _fill(theme=5, tint=0.0)

    # Panel izquierdo: sección (nivel 2) + 2 productos.
    ws.merge_cells("A3:D3")
    ws["A3"] = "Seccion Izq"
    ws["A3"].font = Font(bold=True, size=16)
    ws["A3"].fill = _fill(theme=5, tint=0.4)

    ws.merge_cells("A4:B4")
    ws["A4"] = "Producto Uno"
    ws.merge_cells("A5:B5")
    ws["A5"] = "Producto Dos"

    # Panel derecho: sección (nivel 2) + 1 producto.
    ws.merge_cells("E3:H3")
    ws["E3"] = "Seccion Der"
    ws["E3"].font = Font(bold=True, size=16)
    ws["E3"].fill = _fill(theme=5, tint=0.4)

    ws.merge_cells("E4:F4")
    ws["E4"] = "Producto Tres"

    wb.save(ruta)


def test_parsear_plantilla_ignora_titulo_de_hoja_completa(tmp_path):
    ruta = tmp_path / "plantilla.xlsx"
    _escribir_plantilla_sintetica(str(ruta))
    nodos = parsear_plantilla(str(ruta))
    textos = [n.texto for n in nodos]
    assert "MarcaTest" not in textos


def test_parsear_plantilla_clasifica_encabezado_vs_producto(tmp_path):
    ruta = tmp_path / "plantilla.xlsx"
    _escribir_plantilla_sintetica(str(ruta))
    nodos = parsear_plantilla(str(ruta))

    seccion_izq = next(n for n in nodos if n.texto == "Seccion Izq")
    assert seccion_izq.tipo == TIPO_ENCABEZADO
    assert seccion_izq.panel == "izq"
    assert seccion_izq.nivel == 2

    prod = next(n for n in nodos if n.texto == "Producto Uno")
    assert prod.tipo == TIPO_PRODUCTO
    assert prod.nivel is None
    assert prod.panel == "izq"


def test_parsear_plantilla_paneles_son_independientes(tmp_path):
    ruta = tmp_path / "plantilla.xlsx"
    _escribir_plantilla_sintetica(str(ruta))
    nodos = parsear_plantilla(str(ruta))

    izq = [n for n in nodos if n.panel == "izq"]
    der = [n for n in nodos if n.panel == "der"]
    assert [n.texto for n in izq] == ["Seccion Izq", "Producto Uno", "Producto Dos"]
    assert [n.texto for n in der] == ["Seccion Der", "Producto Tres"]
    # El orden se cuenta independiente por panel, ambos empiezan en 1.
    assert izq[0].orden == 1
    assert der[0].orden == 1


# --- resolver_vinculos: matching conservador por nombre normalizado ---

def _nodo(hoja, panel, orden, tipo, nivel, texto):
    return NodoPlantilla(hoja=hoja, panel=panel, orden=orden, tipo=tipo, nivel=nivel, texto=texto)


def test_resolver_vinculos_matchea_por_subconjunto_de_palabras():
    nodos = [_nodo("Hoja1", "izq", 1, TIPO_PRODUCTO, None, "Pavo Inicio")]
    productos = [{"id": 1, "nombre": "PAVO INICIO 25 KG", "sku": None}]
    assert resolver_vinculos(nodos, productos) == [1]


def test_resolver_vinculos_no_matchea_especie_distinta_por_palabra_generica():
    # "Desarrollo" (bajo Cerdos) NO debe engancharse a "POLLAS DESARROLLO" solo
    # porque comparten la palabra "Desarrollo" — es un cruce de especie real
    # que el resolver debe rechazar (ver PLAN_FASE16.md / CHANGELOG Fase 16).
    nodos = [
        _nodo("Hoja1", "izq", 1, TIPO_ENCABEZADO, 2, "Cerdos"),
        _nodo("Hoja1", "izq", 2, TIPO_PRODUCTO, None, "Desarrollo"),
    ]
    productos = [{"id": 1, "nombre": "POLLAS DESARROLLO 25 kg", "sku": None}]
    assert resolver_vinculos(nodos, productos) == [None, None]


def test_resolver_vinculos_usa_contexto_para_desambiguar_lineas():
    nodos = [
        _nodo("Hoja1", "izq", 1, TIPO_ENCABEZADO, 2, "Cerdos"),
        _nodo("Hoja1", "izq", 2, TIPO_ENCABEZADO, 3, "Linea suprema"),
        _nodo("Hoja1", "izq", 3, TIPO_PRODUCTO, None, "Crecimiento"),
        _nodo("Hoja1", "izq", 4, TIPO_ENCABEZADO, 3, "Linea Optima"),
        _nodo("Hoja1", "izq", 5, TIPO_PRODUCTO, None, "Crecimiento"),
    ]
    productos = [
        {"id": 13, "nombre": "CERDO SUPREMA CRECIMIENTO 25KG", "sku": "s1"},
        {"id": 16, "nombre": "CERDO OPTIMA CRECIMIENTO 25 kg", "sku": "s2"},
    ]
    resultado = resolver_vinculos(nodos, productos)
    assert resultado == [None, None, 13, None, 16]


def test_resolver_vinculos_prefiere_duplicado_con_sku():
    nodos = [_nodo("Hoja1", "izq", 1, TIPO_PRODUCTO, None, "Porcimas")]
    productos = [
        {"id": 27, "nombre": "PORCIMAS 25 kg", "sku": "868"},
        {"id": 99, "nombre": "PORCIMAS 25 kg", "sku": None},
    ]
    assert resolver_vinculos(nodos, productos) == [27]


def test_resolver_vinculos_tamano_explicito_debe_coincidir():
    # La plantilla trae "5kg" -> solo debe matchear el producto de 5kg, no el
    # de 25kg (y viceversa con el renglón sin tamaño explícito) — mismo caso
    # real que "Gallo Regio" / "Gallo Regio 5kg" en Mi patio.
    nodos = [
        _nodo("Hoja1", "der", 1, TIPO_PRODUCTO, None, "Gallo Regio"),
        _nodo("Hoja1", "der", 2, TIPO_PRODUCTO, None, "Gallo Regio 5kg"),
    ]
    productos = [
        {"id": 32, "nombre": "GALLO REGIO 25 kg", "sku": None},
        {"id": 40, "nombre": "GALLO REGIO 5KG", "sku": None},
    ]
    assert resolver_vinculos(nodos, productos) == [32, 40]


def test_resolver_vinculos_palabra_extra_sin_contexto_no_matchea():
    # "CP" no aparece en la plantilla ni en ningún encabezado ancestro -> el
    # resolver NO adivina (queda sin vincular, se resuelve a mano vía el CRUD
    # de edición — ver CHANGELOG "Fase 16").
    nodos = [_nodo("Hoja1", "izq", 1, TIPO_PRODUCTO, None, "Pollo Engorda")]
    productos = [{"id": 25, "nombre": "POLLO ENGORDA CP 25 kg", "sku": None}]
    assert resolver_vinculos(nodos, productos) == [None]


def test_resolver_vinculos_sin_candidatos_queda_sin_vincular():
    nodos = [_nodo("Hoja1", "izq", 1, TIPO_PRODUCTO, None, "Producto Inexistente")]
    assert resolver_vinculos(nodos, []) == [None]


def test_resolver_vinculos_no_reutiliza_el_mismo_producto():
    nodos = [
        _nodo("Hoja1", "izq", 1, TIPO_PRODUCTO, None, "Pavo Inicio"),
        _nodo("Hoja1", "der", 1, TIPO_PRODUCTO, None, "Pavo Inicio"),
    ]
    productos = [{"id": 1, "nombre": "PAVO INICIO 25 KG", "sku": None}]
    resultado = resolver_vinculos(nodos, productos)
    # Solo uno de los dos renglones se queda con el único producto real.
    assert resultado.count(1) == 1
